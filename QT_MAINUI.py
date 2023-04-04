# -*- coding: utf-8 -*-

################################################################################
## Form generated from reading UI file 'QT_2.ui'
##
## Created by: Qt User Interface Compiler version 6.2.3
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################
import os
import csv
from openpyxl import load_workbook
import pandas as pd

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QAction, QBrush, QColor, QConicalGradient,
    QCursor, QFont, QFontDatabase, QGradient,
    QIcon, QImage, QKeySequence, QLinearGradient,
    QPainter, QPalette, QPixmap, QRadialGradient,QScreen,
    QTransform)
from PySide6.QtWidgets import (QApplication, QHBoxLayout, QHeaderView, QLabel,
    QMainWindow, QMenu, QMenuBar, QPushButton, QDialog,QTabWidget,QTableView,
    QSizePolicy, QSpacerItem, QStatusBar,QFileDialog,QMessageBox, 
    QTableWidget, QTableWidgetItem, QTextBrowser, QWidget)
from Parameter_ui import Ui_Dialog_Para
from Tooltip_ui import Ui_Dialog_Tool
from User_manage import Ui_Dialog_User

class Ui_MainWindow(object):
    def setupUi(self, MainWindow, isadmin):
        self.parent = MainWindow
        self.path_Need_name = None
        self.path_Roll_name = None
        self.Need_file_encoding = None
        self.Roll_file_encoding = None
        self.plan = None
        self.roll_length = None
        self.Roll_file_new = None
        self.final_info = [] #最后一行信息
        if not MainWindow.objectName():
            MainWindow.setObjectName(u"MainWindow")
        MainWindow.resize(1399, 795)
        MainWindow.setMinimumSize(QSize(1399, 795))
        MainWindow.setMaximumSize(QSize(1399, 795))
        center = QScreen.availableGeometry(QApplication.primaryScreen()).center()
        geo = MainWindow.frameGeometry()
        geo.moveCenter(center)
        MainWindow.move(geo.topLeft())
        font = QFont()
        font.setFamilies([u"\u5b8b\u4f53"])
        font.setPointSize(14)
        font.setBold(False)
        MainWindow.setFont(font)
        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName(u"centralwidget")
        font1 = QFont()
        font1.setFamilies([u"\u5b8b\u4f53"])
        font1.setPointSize(12)
        font1.setBold(False)
        # self.tableWidget_5 = QTableWidget(self.centralwidget)
        # self.tableWidget_5.setObjectName(u"tableWidget_5")
        # self.tableWidget_5.setGeometry(QRect(40, 450, 1321, 231))
        # self.tableWidget_5.setFont(font1)
        self.tabWidget = QTabWidget(self.centralwidget)
        self.tabWidget.setObjectName(u"tabWidget")
        self.tabWidget.setGeometry(QRect(40, 430, 1321, 261))
        font3 = QFont()
        font3.setFamilies([u"\u5b8b\u4f53"])
        font3.setPointSize(16)
        self.tabWidget.setFont(font3)
        self.tab = QWidget()
        self.tab.setObjectName(u"tab")
        self.tableWidget_5 = QTableWidget(self.tab)
        self.tableWidget_5.setObjectName(u"tableView_2")
        self.tableWidget_5.setGeometry(QRect(0, -3, 1321, 231))
        self.tableWidget_5.setFont(font1)
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QWidget()
        self.tab_2.setObjectName(u"tab_2")
        self.tableWidget_6 = QTableWidget(self.tab_2)
        self.tableWidget_6.setObjectName(u"tableWidget")
        self.tableWidget_6.setGeometry(QRect(0, -3, 1321, 231))
        self.tableWidget_6.setFont(font1)
        self.tabWidget.addTab(self.tab_2, "")

        self.pushButton_4 = QPushButton(self.centralwidget)
        self.pushButton_4.setObjectName(u"pushButton_4")
        self.pushButton_4.setGeometry(QRect(280, 360, 201, 51))
        self.pushButton_4.clicked.connect(self.show_Tooltip)
        font2 = QFont()
        font2.setFamilies([u"\u5b8b\u4f53"])
        font2.setPointSize(16)
        font2.setBold(False)
        self.pushButton_4.setFont(font2)
        self.pushButton_5 = QPushButton(self.centralwidget)
        self.pushButton_5.setObjectName(u"pushButton_5")
        self.pushButton_5.setGeometry(QRect(950, 360, 201, 51))
        self.pushButton_5.clicked.connect(self.save_result)
        self.pushButton_5.setFont(font2)
        
        self.layoutWidget = QWidget(self.centralwidget)
        self.layoutWidget.setObjectName(u"layoutWidget")
        self.layoutWidget.setGeometry(QRect(34, 80, 691, 261))
        self.layoutWidget.setFont(font1)
        self.horizontalLayout = QHBoxLayout(self.layoutWidget)
        self.horizontalLayout.setSpacing(2)
        self.horizontalLayout.setObjectName(u"horizontalLayout")
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.tableWidget = QTableWidget(self.layoutWidget)
        self.tableWidget.setObjectName(u"tableWidget")

        self.horizontalLayout.addWidget(self.tableWidget)

        self.tableWidget_3 = QTableWidget(self.layoutWidget)
        self.tableWidget_3.setObjectName(u"tableWidget_3")
        self.tableWidget_3.setFont(font1)

        self.horizontalLayout.addWidget(self.tableWidget_3)

        self.tableWidget_2 = QTableWidget(self.layoutWidget)
        self.tableWidget_2.setObjectName(u"tableWidget_2")
        self.tableWidget_2.setFont(font1)

        self.horizontalLayout.addWidget(self.tableWidget_2)

        self.layoutWidget1 = QWidget(self.centralwidget)
        self.layoutWidget1.setObjectName(u"layoutWidget1")
        self.layoutWidget1.setGeometry(QRect(40, 410, 1381, 31))
        self.horizontalLayout_6 = QHBoxLayout(self.layoutWidget1)
        self.horizontalLayout_6.setObjectName(u"horizontalLayout_6")
        self.horizontalLayout_6.setContentsMargins(0, 0, 0, 0)
        # self.label_4 = QLabel(self.layoutWidget1)
        # self.label_4.setObjectName(u"label_4")
        # self.label_4.setFont(font2)

        # self.horizontalLayout_6.addWidget(self.label_4)

        self.label_5 = QLabel(self.layoutWidget1)
        self.label_5.setObjectName(u"label_5")
        self.label_5.setFont(font2)

        self.horizontalLayout_6.addWidget(self.label_5)

        self.layoutWidget2 = QWidget(self.centralwidget)
        self.layoutWidget2.setObjectName(u"layoutWidget2")
        self.layoutWidget2.setGeometry(QRect(42, 712, 1381, 31))
        self.horizontalLayout_7 = QHBoxLayout(self.layoutWidget2)
        self.horizontalLayout_7.setObjectName(u"horizontalLayout_7")
        self.horizontalLayout_7.setContentsMargins(0, 0, 0, 0)
        self.label_6 = QLabel(self.layoutWidget2)
        self.label_6.setObjectName(u"label_6")
        self.label_6.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_6)

        self.label_8 = QLabel(self.layoutWidget2)
        self.label_8.setObjectName(u"label_8")
        self.label_8.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_8)

        self.label_10 = QLabel(self.layoutWidget2)
        self.label_10.setObjectName(u"label_10")
        self.label_10.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_10)

        self.label_12 = QLabel(self.layoutWidget2)
        self.label_12.setObjectName(u"label_12")
        self.label_12.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_12)

        self.label_7 = QLabel(self.layoutWidget2)
        self.label_7.setObjectName(u"label_7")
        self.label_7.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_7)

        self.label_11 = QLabel(self.layoutWidget2)
        self.label_11.setObjectName(u"label_11")
        self.label_11.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_11)

        self.label_13 = QLabel(self.layoutWidget2)
        self.label_13.setObjectName(u"label_13")
        self.label_13.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_13)

        self.label_14 = QLabel(self.layoutWidget2)
        self.label_14.setObjectName(u"label_14")
        self.label_14.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_14)

        self.label_15 = QLabel(self.layoutWidget2)
        self.label_15.setObjectName(u"label_15")
        self.label_15.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_15)

        self.label_16 = QLabel(self.layoutWidget2)
        self.label_16.setObjectName(u"label_16")
        self.label_16.setFont(font2)

        self.horizontalLayout_7.addWidget(self.label_16)

        self.tableWidget_4 = QTableWidget(self.centralwidget)
        self.tableWidget_4.setObjectName(u"tableWidget_4")
        self.tableWidget_4.setGeometry(QRect(740, 80, 621, 261))
        self.tableWidget_4.setFont(font1)
        self.label_3 = QLabel(self.centralwidget)
        self.label_3.setObjectName(u"label_3")
        self.label_3.setGeometry(QRect(693, 31, 16, 22))
        self.widget = QWidget(self.centralwidget)
        self.widget.setObjectName(u"widget")
        self.widget.setGeometry(QRect(35, 31, 691, 41))
        self.horizontalLayout_3 = QHBoxLayout(self.widget)
        self.horizontalLayout_3.setObjectName(u"horizontalLayout_3")
        self.horizontalLayout_3.setContentsMargins(0, 0, 0, 0)
        self.label = QLabel(self.widget)
        self.label.setObjectName(u"label")
        self.label.setFont(font2)

        self.horizontalLayout_3.addWidget(self.label)

        self.horizontalSpacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)

        self.horizontalLayout_3.addItem(self.horizontalSpacer)

        self.pushButton = QPushButton(self.widget)
        self.pushButton.setObjectName(u"pushButton")
        self.pushButton.setFont(font2)
        self.pushButton.clicked.connect(self.Open_Need_File)
        self.horizontalLayout_3.addWidget(self.pushButton)

        self.widget1 = QWidget(self.centralwidget)
        self.widget1.setObjectName(u"widget1")
        self.widget1.setGeometry(QRect(740, 30, 621, 41))
        self.horizontalLayout_4 = QHBoxLayout(self.widget1)
        self.horizontalLayout_4.setObjectName(u"horizontalLayout_4")
        self.horizontalLayout_4.setContentsMargins(0, 0, 0, 0)
        self.label_2 = QLabel(self.widget1)
        self.label_2.setObjectName(u"label_2")
        self.label_2.setFont(font2)

        self.horizontalLayout_4.addWidget(self.label_2)

        self.horizontalSpacer_2 = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)

        self.horizontalLayout_4.addItem(self.horizontalSpacer_2)

        self.pushButton_2 = QPushButton(self.widget1)
        self.pushButton_2.setObjectName(u"pushButton_2")
        self.pushButton_2.setFont(font2)

        self.horizontalLayout_4.addWidget(self.pushButton_2)
        
        # 原材料导入按钮
        self.pushButton_2.clicked.connect(self.Open_Roll_File)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QMenuBar(MainWindow)
        self.menubar.setObjectName(u"menubar")
        self.menubar.setGeometry(QRect(0, 0, 1341, 25))
        self.menu = QMenu(self.menubar)
        self.menu.setContextMenuPolicy(Qt.NoContextMenu)
        self.menu.setToolTipDuration(5)
        self.menu.setToolTipsVisible(False)
        self.menu.setObjectName(u"menu")
        
        self.setAction = QAction(self.parent)
        self.setAction.setCheckable(False)
        self.setAction.setObjectName('setAction')
        self.setAction.triggered.connect(self.show_para_ui)
        self.setAction.setText('参数配置')
        self.menubar.addAction(self.setAction)
        if isadmin == True:
            self.setAction_2 = QAction(self.parent)
            self.setAction_2.setCheckable(False)
            self.setAction_2.setObjectName('setAction_2')
            self.setAction_2.triggered.connect(self.show_user_ui)
            self.setAction_2.setText('用户管理')
            self.menubar.addAction(self.setAction_2)

            self.setAction_3 = QAction(self.parent)
            self.setAction_3.setCheckable(False)
            self.setAction_3.setObjectName('setAction_3')
            str2 = ''
            for i in range(106-len(self.parent.name.encode('utf-8'))):
                str2 = str2 + ' '
            self.setAction_3.setText(str2)
            self.menubar.addAction(self.setAction_3)
            self.setAction_4 = QAction(self.parent)
            self.setAction_4.setCheckable(False)
            self.setAction_4.setObjectName('setAction_4')
            str3 = '欢迎' + self.parent.name + '登录' 
            self.setAction_4.setText(str3)
            self.menubar.addAction(self.setAction_4)

            self.setAction_5 = QAction(self.parent)
            self.setAction_5.setCheckable(False)
            self.setAction_5.setObjectName('setAction_5')
            self.setAction_5.triggered.connect(self.login_out)
            self.setAction_5.setText('退出登录')
            self.menubar.addAction(self.setAction_5)
        elif isadmin == False:
            self.setAction_2 = QAction(self.parent)
            self.setAction_2.setCheckable(False)
            self.setAction_2.setObjectName('setAction_2')
            str2 = ''
            for i in range(116-len(self.parent.name.encode('utf-8'))):
                str2 = str2 + ' '
            self.setAction_2.setText(str2)
            self.menubar.addAction(self.setAction_2)
            self.setAction_3 = QAction(self.parent)
            self.setAction_3.setCheckable(False)
            self.setAction_3.setObjectName('setAction_3')
            str1 = '欢迎' + self.parent.name + '登录' 
            self.setAction_3.setText(str1)
            self.menubar.addAction(self.setAction_3)

            self.setAction_4 = QAction(self.parent)
            self.setAction_4.setCheckable(False)
            self.setAction_4.setObjectName('setAction_4')
            self.setAction_4.triggered.connect(self.login_out)
            self.setAction_4.setText('退出登录')
            self.menubar.addAction(self.setAction_4)

        MainWindow.setMenuBar(self.menubar)

        self.statusbar = QStatusBar(MainWindow)
        self.statusbar.setObjectName(u"statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.menubar.addAction(self.menu.menuAction())
        self.tabWidget.setCurrentIndex(0)
        self.retranslateUi(MainWindow)
        QMetaObject.connectSlotsByName(MainWindow)
    # setupUi
    def show_para_ui(self):
        self.Para = Ui_Dialog_Para(self.parent)
        self.Para.exec()

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(QCoreApplication.translate("MainWindow", u"\u4e0a\u6d77\u4e2d\u6d66\u53d6\u5411\u7845\u94a2\u667a\u80fd\u914d\u6599\u7cfb\u7edf", None))
        self.pushButton_4.setText(QCoreApplication.translate("MainWindow", u"\u5f00\u59cb\u914d\u6599", None))
        self.pushButton_5.setText(QCoreApplication.translate("MainWindow", u"\u5bfc\u51fa\u7ed3\u679c", None))
        self.label.setText(QCoreApplication.translate("MainWindow", u"\u5ba2\u6237\u9700\u6c42\u4fe1\u606f\uff1a", None))
        self.pushButton.setText(QCoreApplication.translate("MainWindow", u"\u5bfc\u5165", None))
        self.label_3.setText("")
        self.label_2.setText(QCoreApplication.translate("MainWindow", u"\u539f\u6750\u6599\u4fe1\u606f\uff1a", None))
        self.pushButton_2.setText(QCoreApplication.translate("MainWindow", u"\u5bfc\u5165", None))
        # self.label_4.setText(QCoreApplication.translate("MainWindow", u"\u539f\u6750\u6599\u5207\u5272\u65b9\u6848\uff1a", None))
        # self.label_5.setText(QCoreApplication.translate("MainWindow", u"\u9700\u6c42\u5206\u914d\u65b9\u6848\uff1a", None))
        self.label_6.setText(QCoreApplication.translate("MainWindow", u"\u4e3b\u9879\u76ee\u7387\uff1a", None))
        # self.label_8.setText(QCoreApplication.translate("MainWindow", u"%", None))
        self.label_10.setText(QCoreApplication.translate("MainWindow", u"\u6b21\u9879\u76ee\u7387\uff1a", None))
        # self.label_12.setText(QCoreApplication.translate("MainWindow", u"%", None))
        self.label_7.setText(QCoreApplication.translate("MainWindow", u"\u7b2c\u4e09\u9879\u76ee\u7387\uff1a", None))
        # self.label_11.setText(QCoreApplication.translate("MainWindow", u"%", None))
        self.label_13.setText(QCoreApplication.translate("MainWindow", u"\u603b\u9879\u76ee\u7387\uff1a", None))
        self.label_15.setText(QCoreApplication.translate("MainWindow", u"\u603b\u914d\u5200\u6570\uff1a", None))
        # self.label_14.setText(QCoreApplication.translate("MainWindow", u"%", None))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), QCoreApplication.translate("MainWindow", u"\u914d\u6599\u65b9\u6848", None))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), QCoreApplication.translate("MainWindow", u"\u9700\u6c42\u6ee1\u8db3\u60c5\u51b5", None))
    # retranslateUi
    def cal_length(self,width,weight,para):
        return weight/width/7.65/para*1000

    def Get_csv_encoding(self, path_Roll_name):
        encodings = ['utf-8-sig','utf-8','gbk','GB2312','gb18030',]
        for e in encodings:
            try:
                with open(path_Roll_name, 'r', encoding=e) as f:
                    reader = csv.reader(f)
                    for item in reader:
                        a = item
                print("编码为：", e)
                break
            except:
                print("不适用编码为：", e)
        return e

    def Open_Need_File(self):
        Need_name = QFileDialog.getOpenFileName(self.parent, '选择文件', '', 'Excel files(*.xlsx *.csv *.CSV)')  # 过滤出xlsx格式的文件
        self.path_Need_name = Need_name[0]
        print("需求文档的路径：", self.path_Need_name)
        if len(self.path_Need_name) > 0:
            if self.path_Need_name.split("/")[-1].split(".")[-1] == "csv":
                self.FILE_CSV = True
            else:
                self.FILE_CSV = False
            self.Read_Need_Data()
            self.Need_Table_Initialize(0.285)

    # 读取 需求 文档的数据
    def Read_Need_Data(self):
        if self.FILE_CSV:
            self.Need_file_encoding = self.Get_csv_encoding(self.path_Need_name)
            self.Need_file = pd.read_csv(self.path_Need_name,encoding=self.Need_file_encoding, header=None)
            self.Need_file.drop(self.Need_file.columns[[2,]],axis=1,inplace=True)
            print(self.Need_file)
            # 获取数据的行标签与列标签
        else:
            self.Need_file = pd.read_excel(self.path_Need_name,header=None)
            self.Need_file.drop(self.Need_file.columns[[2,]],axis=1,inplace=True)
            print(self.Need_file)
            # 获取数据的行标签与列标签

        # 定义表格的初始化参数
    def Need_Table_Initialize(self,para):
        self.project_index = []
        self.rows = len(self.Need_file.index)
        self.columns = len(self.Need_file.columns)
        print("columns:"+str(self.columns))
        self.project_index = []
        flag_wukehu = False
        for i in range(self.rows):
            tstr = self.Need_file.iloc[i][0]
            if str(tstr).isdigit() == False:
                if str(tstr).split('）')[0] == str(tstr):
                    if len(str(tstr).strip().split('(')[0]) >= 3:
                        if str(tstr).strip().split('(')[0][:1] == '无':
                            flag_wukehu = True
                        self.project_index.append(i)
                else:
                    if len(str(tstr).strip().split('（')[0]) >= 3:
                        if str(tstr).strip().split('（')[0][:1] == '无':
                            flag_wukehu = True
                        self.project_index.append(i)
        if flag_wukehu == False:                
            for i in range(self.rows):
                tstr = self.Need_file.iloc[i][0]
                if str(tstr).isdigit() == False:
                    if str(tstr).split('）')[0] == str(tstr):
                        if str(tstr).strip().split('(')[0] == '边条':
                            self.project_index.append(i)
                            break
                    else:
                        if str(tstr).strip().split('（')[0] == '边条':
                            self.project_index.append(i)
                            break
        # print(flag_wukehu)
        # print(self.project_index)
        list_with_diff = []
        for n in range(1, len(self.project_index)):
            diff = self.project_index[n] - self.project_index[n-1]
            if diff != 1:
                list_with_diff.append(diff)

        self.tableWidget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget.verticalHeader().setHidden(True)
        self.tableWidget.horizontalHeader().setHidden(True)
        self.tableWidget_3.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget_3.verticalHeader().setHidden(True)
        self.tableWidget_3.horizontalHeader().setHidden(True)
        self.tableWidget_2.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget_2.verticalHeader().setHidden(True)
        self.tableWidget_2.horizontalHeader().setHidden(True)

        row_three = [0,0,0]
        for i in range(len(list_with_diff)):
            index = i % 3
            row_three[index] = row_three[index] + list_with_diff[i]
        
        self.tableWidget.setRowCount(row_three[0])    # 设置表格的行数
        self.tableWidget.setColumnCount(self.columns + 2)  # 设置表格的列数
        self.tableWidget_3.setRowCount(row_three[1])    # 设置表格的行数
        self.tableWidget_3.setColumnCount(self.columns + 2)  # 设置表格的列数
        self.tableWidget_2.setRowCount(row_three[2])    # 设置表格的行数
        self.tableWidget_2.setColumnCount(self.columns + 2)  # 设置表格的列数
        item_index = []
        for i in range(len(list_with_diff)):
            if i / 3 == 0:
                item_index.append(0)
            else:
                sum1 = 0
                for j in range(int(i/3)):
                    sum1 = sum1 + list_with_diff[j*3+i%3]
                item_index.append(sum1)
        mid = []
        length = []
        for i in range(len(list_with_diff)):
            mid.append('')
            mid.append('id')
            length.append('')
            length.append('长度')
            for j in range(list_with_diff[i] - 2):
                mid.append(j+1)
            for j in range(list_with_diff[i]):
                if str(self.Need_file.loc[j + sum(list_with_diff[:i])][0]).isdigit() == True:
                    length.append(round(self.cal_length(int(self.Need_file.loc[j + sum(list_with_diff[:i])][0]),int(self.Need_file.loc[j + sum(list_with_diff[:i])][1]),para)))

        for i in range(len(list_with_diff)):
            if (i % 3) == 0:
                for m in range(list_with_diff[i]):
                    if m != 0:
                        index1 = sum(list_with_diff[:i]) + m
                        str2 = str(mid[index1])
                        data_id = QTableWidgetItem(str2)
                        self.tableWidget.setItem(item_index[i] + m, 0, data_id)
                        self.tableWidget.item(item_index[i] + m, 0).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                        str3 = str(length[index1])
                        data_id2 = QTableWidgetItem(str3)
                        self.tableWidget.setItem(item_index[i] + m, 3, data_id2)
                        self.tableWidget.item(item_index[i] + m, 3).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                    for n in range(self.columns):
                        index1 = sum(list_with_diff[:i]) + m
                        tstr = str(self.Need_file.iloc[index1][n])
                        data = QTableWidgetItem(tstr)
                        if m == 0 and m == 0:
                            self.tableWidget.setItem(item_index[i] + m, n, data)
                            self.tableWidget.item(item_index[i] + m, n).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                        else:
                            self.tableWidget.setItem(item_index[i] + m, n + 1, data)
                            self.tableWidget.item(item_index[i] + m, n + 1).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                self.tableWidget.setSpan(item_index[i], 0, 1, 4)
            elif (i % 3) == 1:
                for m in range(list_with_diff[i]):
                    if m != 0:
                        index2 = sum(list_with_diff[:i]) + m
                        str2 = str(mid[index2])
                        data_id = QTableWidgetItem(str2)
                        self.tableWidget_3.setItem(item_index[i] + m, 0, data_id)
                        self.tableWidget_3.item(item_index[i] + m, 0).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                        str3 = str(length[index2])
                        data_id2 = QTableWidgetItem(str3)
                        self.tableWidget_3.setItem(item_index[i] + m, 3, data_id2)
                        self.tableWidget_3.item(item_index[i] + m, 3).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                    for n in range(self.columns):
                        index2 = sum(list_with_diff[:i]) + m
                        tstr = str(self.Need_file.iloc[index2][n])
                        data = QTableWidgetItem(tstr)
                        if m == 0 and m == 0:
                            self.tableWidget_3.setItem(item_index[i] + m, n, data)
                            self.tableWidget_3.item(item_index[i] + m, n).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                        else:
                            self.tableWidget_3.setItem(item_index[i] + m, n + 1, data)
                            self.tableWidget_3.item(item_index[i] + m, n + 1).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                self.tableWidget_3.setSpan(item_index[i], 0, 1, 4)
            else:
                for m in range(list_with_diff[i]):
                    if m != 0:
                        index3 = sum(list_with_diff[:i]) + m
                        str2 = str(mid[index3])
                        data_id = QTableWidgetItem(str2)
                        self.tableWidget_2.setItem(item_index[i] + m, 0, data_id)
                        self.tableWidget_2.item(item_index[i] + m, 0).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                        str3 = str(length[index3])
                        data_id2 = QTableWidgetItem(str3)
                        self.tableWidget_2.setItem(item_index[i] + m, 3, data_id2)
                        self.tableWidget_2.item(item_index[i] + m, 3).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                    for n in range(self.columns):
                        index3 = sum(list_with_diff[:i]) + m
                        tstr = str(self.Need_file.iloc[index3][n])
                        data = QTableWidgetItem(tstr)
                        if m == 0 and m == 0:
                            self.tableWidget_2.setItem(item_index[i] + m, n, data)
                            self.tableWidget_2.item(item_index[i] + m, n).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                        else:
                            self.tableWidget_2.setItem(item_index[i] + m, n + 1, data)
                            self.tableWidget_2.item(item_index[i] + m, n + 1).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
                self.tableWidget_2.setSpan(item_index[i], 0, 1, 4)
                        
        self.tableWidget.resizeColumnsToContents() # 列宽随着内容调整
        self.tableWidget.resizeRowsToContents() # 行宽随着内容调整
        self.tableWidget.setAlternatingRowColors(True)  # 表格的颜色交错显示
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_3.resizeColumnsToContents() # 列宽随着内容调整
        self.tableWidget_3.resizeRowsToContents() # 行宽随着内容调整
        self.tableWidget_3.setAlternatingRowColors(True)  # 表格的颜色交错显示
        self.tableWidget_3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_2.resizeColumnsToContents() # 列宽随着内容调整
        self.tableWidget_2.resizeRowsToContents() # 行宽随着内容调整
        self.tableWidget_2.setAlternatingRowColors(True)  # 表格的颜色交错显示
        self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for i in range(row_three[0]):
            self.tableWidget.setRowHeight(i, int(self.tableWidget_2.height() * 0.1))
        for i in range(row_three[1]):
            self.tableWidget_3.setRowHeight(i, int(self.tableWidget_2.height() * 0.1))
        for i in range(row_three[2]):
            self.tableWidget_2.setRowHeight(i, int(self.tableWidget_2.height() * 0.1))

    def Open_Roll_File(self):
        Roll_name = QFileDialog.getOpenFileName(self.parent, '选择文件', '', 'Excel files(*.xlsx *.csv *.CSV)')   # 过滤出xlsx格式的文件
        self.path_Roll_name = Roll_name[0]
        print("需求文档的路径：", self.path_Roll_name)
        if len(self.path_Roll_name) > 0:
            if self.path_Roll_name.split("/")[-1].split(".")[-1] == "csv":
                self.FILE_CSV = True
            else:
                self.FILE_CSV = False
            self.Read_Roll_Data()
            self.Roll_Table_Initialize()

    # 读取 需求 文档的数据
    def Read_Roll_Data(self):
        if self.FILE_CSV:
            self.Roll_file_encoding = self.Get_csv_encoding(self.path_Roll_name)
            self.Roll_file = pd.read_csv(self.path_Roll_name,encoding=self.Roll_file_encoding)
            self.Roll_file = self.Roll_file.fillna('')
        else:
            self.Roll_file = pd.read_excel(self.path_Roll_name)
            self.Roll_file = self.Roll_file.fillna('')

        # 定义表格的初始化参数
    def Roll_Table_Initialize(self):
        mid = self.Roll_file['编号']
        self.Roll_file.drop(labels=['编号'], axis=1,inplace = True)
        self.Roll_file.insert(2, '编号', mid)
        new_mid = [i for i in range(1,len(mid.index)+1)]
        self.Roll_file.insert(0, 'id', new_mid)
        header_list = ['id', '卷性质','牌号','捆包号/卷号','宽度','重量','铁损','磁感']
        self.Roll_file.columns = header_list
        length = []
        self.rows = self.Roll_file.index
        for i in range(len(self.rows)):
            length.append(round(self.cal_length(int(self.Roll_file.iloc[i][4]),int(self.Roll_file.iloc[i][5]),0.285)))
        self.Roll_file['长度'] = pd.Series(length)
        self.columns = self.Roll_file.columns

        self.tableWidget_4.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget_4.setRowCount(len(self.rows))    # 设置表格的行数
        self.tableWidget_4.setColumnCount(len(self.columns))  # 设置表格的列数
        self.tableWidget_4.setHorizontalHeaderLabels(self.columns) # 设置表格的列标签
        self.tableWidget_4.verticalHeader().setHidden(True)
        for i in range(len(self.rows)):
            for j in range(len(self.columns)):
                tstr = str(self.Roll_file.iloc[i][j])
                data = QTableWidgetItem(tstr)
                self.tableWidget_4.setItem(i, j, data)
                self.tableWidget_4.item(i, j).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
        self.tableWidget_4.horizontalHeader().setStyleSheet("border-top:1px solid rgb(210, 210, 210);border-bottom:1px solid rgb(210, 210, 210);font:12pt '宋体';")
        self.tableWidget_4.resizeColumnsToContents() # 列宽随着内容调整
        self.tableWidget_4.resizeRowsToContents() # 行宽随着内容调整
        self.tableWidget_4.setAlternatingRowColors(True)  # 表格的颜色交错显示
        for i in range(len(self.columns)):
            if i == 3:
                self.tableWidget_4.setColumnWidth(i, int(self.tableWidget_4.width() * 0.2))
            elif i == 0:
                self.tableWidget_4.setColumnWidth(i, int(self.tableWidget_4.width() * 0.1))
            else:
                self.tableWidget_4.setColumnWidth(i, int(self.tableWidget_4.width() * 0.12))
        # self.tableWidget_4.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def show_Tooltip(self):
        #Tool = Ui_Dialog_Tool()
        self.tableWidget_5.clear()
        self.tableWidget_6.clear()
        self.label_8.clear()
        self.label_12.clear()
        self.label_11.clear()
        self.label_14.clear()
        self.label_16.clear()
        if self.parent.progress.isVisible():
            self.parent.progress.stop()
        self.Tool = Ui_Dialog_Tool(self.parent, self.path_Need_name, self.path_Roll_name, self.Need_file_encoding, self.Roll_file_encoding)
        self.Tool.exec()
    
    def setALL(self,plan,prolist,list_rate,all_rate, main_len, roll_length,program_ok,idlist,leftcut):
        # print(prolist)
        self.plan = plan
        self.roll_length = roll_length
        if self.parent.Par_data != None:
            self.Need_Table_Initialize(self.parent.Par_data.THICKNESS)

        self.Roll_file['长度'] = pd.Series(self.roll_length)
        rows2 = self.Roll_file.index
        columns2 = self.Roll_file.columns
        for i in range(len(rows2)):
            for j in range(len(columns2)):
                tstr = str(self.Roll_file.iloc[i][j])
                data = QTableWidgetItem(tstr)
                self.tableWidget_4.setItem(i, j, data)
                self.tableWidget_4.item(i, j).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
        
        Roll_file_copy = self.Roll_file.copy()
        if main_len != 1:
            QMessageBox.warning(self.parent, "提示", "主需求没有全部满足", QMessageBox.Ok)
        
        self.final_info = []
        maxlen = - float('inf')
        sum_cut = 0
        print(self.plan.items())
        for item,value in self.plan.items():
            value_len = len(value)
            sum_cut = sum_cut + value_len
            if value_len > maxlen:
                maxlen = value_len
        mid = Roll_file_copy['捆包号/卷号'].astype(str)
        Roll_file_copy.drop(labels=['捆包号/卷号'], axis=1,inplace = True)
        Roll_file_copy.insert(3, '捆包号/卷号', mid)
        header_list = ['id', '卷性质','牌号','捆包号/卷号','宽度','重量','铁损','磁感','长度','退卷']
        for i in range(int(maxlen)):
            header_list.append('配刀'+str(i+1))
            header_list.append('米数'+str(i+1))
            header_list.append('边丝'+str(i+1))
        for i in range(maxlen*3 + 1):
            if i == 0:
                Roll_file_copy[str(i)] = 0
            else:
                Roll_file_copy[str(i)] = ''
        Roll_file_copy.columns = header_list
        Roll_file_copy['长度'] = pd.Series(self.roll_length)
        for item,value in self.plan.items():
            for j in range(len(value)):
                result_list = value[j]['width'].split('^')
                Roll_file_copy.loc[Roll_file_copy['捆包号/卷号'] == str(item), '配刀'+str(j+1)] = result_list[0]
                Roll_file_copy.loc[Roll_file_copy['捆包号/卷号'] == str(item), '米数'+str(j+1)] = round(value[j]['length'])
                Roll_file_copy.loc[Roll_file_copy['捆包号/卷号'] == str(item), '退卷'] = Roll_file_copy.loc[Roll_file_copy['捆包号/卷号'] == str(item), '退卷'] + round(value[j]['length'])
                if len(result_list) >= 2:
                    Roll_file_copy.loc[Roll_file_copy['捆包号/卷号'] == str(item), '边丝'+str(j+1)] = round(float(result_list[1])*2)
        Roll_file_copy['退卷'] = Roll_file_copy['长度'] - Roll_file_copy['退卷']
        # Roll_file_copy['退卷'] = Roll_file_copy.apply(lambda x: x if x['退卷'] != x['长度'] else 0, axis=1)
        def valuation_formula(x, y):
            if x == y:
                return 0
            else:
                return x
        Roll_file_copy['退卷'] = Roll_file_copy.apply(lambda row: valuation_formula(row['退卷'], row['长度']), axis=1)

        self.Roll_file_new = Roll_file_copy.copy()
        
        columns = self.Roll_file_new.columns
        rows = self.Roll_file_new.index

        self.tableWidget_5.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget_5.setRowCount(len(rows))    # 设置表格的行数
        self.tableWidget_5.setColumnCount(len(columns))  # 设置表格的列数
        self.tableWidget_5.setHorizontalHeaderLabels(columns) # 设置表格的列标签
        self.tableWidget_5.verticalHeader().setHidden(True)

        print(idlist)

        for i in range(len(idlist)):
            for j in range(i,len(idlist)):
                if str(idlist[i])==str(self.Roll_file_new.iloc[j]['捆包号/卷号']):
                    temp=self.Roll_file_new.iloc[i]
                    self.Roll_file_new.iloc[i]=self.Roll_file_new.iloc[j]
                    self.Roll_file_new.iloc[j]=temp

        
        # 将 EXCEL 文档的数据显示在表格中
        for i in range(len(rows)):
            for j in range(len(columns)):
                tstr = str(self.Roll_file_new.iloc[i][j])
                data = QTableWidgetItem(tstr)
                self.tableWidget_5.setItem(i, j, data)
                self.tableWidget_5.item(i, j).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
        self.tableWidget_5.horizontalHeader().setStyleSheet("border-top:1px solid rgb(210, 210, 210);border-bottom:1px solid rgb(210, 210, 210);font:12pt '宋体';")
        self.tableWidget_5.resizeColumnsToContents() # 列宽随着内容调整
        self.tableWidget_5.resizeRowsToContents() # 行宽随着内容调整
        self.tableWidget_5.setAlternatingRowColors(True)  # 表格的颜色交错显示
        self.show_need(prolist,self.parent.Par_data.THICKNESS,self.parent.Par_data.MIN_LENGTH,self.parent.Par_data.POSITIVE_ERROR_PERCENT,self.parent.Par_data.NEGATIVE_ERROR_PERCENT)
        # for i in range(len(self.columns)):
        #     if i == 0:
        #         self.tableWidget_4.setColumnWidth(i, int(self.tableWidget_4.width() * 0.2))
        #     else:
        #         self.tableWidget_4.setColumnWidth(i, int(self.tableWidget_4.width() * 0.15))
        # self.tableWidget_5.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.final_info.append('指标')
        self.final_info.append('主项目率')
        self.final_info.append(str(list_rate[0])+'%')
        if len(list_rate) == 1:
            self.label_8.setText(str(list_rate[0])+'%')
        elif(len(list_rate) == 2):
            self.label_8.setText(str(list_rate[0])+'%')
            self.label_12.setText(str(list_rate[1])+'%')
            self.final_info.append('次项目率')
            self.final_info.append(str(list_rate[1])+'%')
        else:
            self.label_8.setText(str(list_rate[0])+'%')
            self.label_12.setText(str(list_rate[1])+'%')
            self.label_11.setText(str(list_rate[2])+'%')
            self.final_info.append('次项目率')
            self.final_info.append(str(list_rate[1])+'%')
            self.final_info.append('第三项目率')
            self.final_info.append(str(list_rate[2])+'%')
        self.label_14.setText(str(all_rate)+'%')
        self.label_16.setText(str(sum_cut-leftcut)+'刀')
        self.final_info.append('总项目率')
        self.final_info.append(str(all_rate)+'%')
        self.final_info.append('总配刀数')
        self.final_info.append(str(sum_cut-leftcut)+'刀')

    def  num_to_char(self,num):
        num_dict={"0":u"零","1":u"一","2":u"二","3":u"三","4":u"四","5":u"五","6":u"六","7":u"七","8":u"八","9":u"九"}
        return num_dict[str(num)]

    def cal_weight(self,sum_length,width, para):
        return sum_length/1000*para*7.65*width

    def show_need(self,prolist,para,min_length,per,ner):
        print(prolist)
        for i in range(len(self.project_index)-1):
            if i == 0:
                Need = self.Need_file[self.project_index[i]:self.project_index[i+1]].copy()
                Need = Need.reset_index(drop=True).drop([0,1]).reset_index(drop=True)
            else:
                Need_2 = self.Need_file[self.project_index[i]:self.project_index[i+1]].copy()
                print(Need_2)
                Need_2 = Need_2.reset_index(drop=True).drop([0,1]).reset_index(drop=True)
                Need = Need.append(Need_2)
        Need = Need.reset_index(drop=True)
        list_with_diff = []
        for n in range(1, len(self.project_index)):
            diff = self.project_index[n] - self.project_index[n-1]
            if diff != 1:
                list_with_diff.append(diff-2)
        # print(list_with_diff)
        list_value = []
        for i in range(len(list_with_diff)):
            for j in range(list_with_diff[i]):
                if i == 0:
                    list_value.append('主项目')
                elif i== 1:
                    list_value.append('次项目')
                else:
                    list_value.append('第' + self.num_to_char(i+1) + '项目')
        # print(list_value)
        mid = []
        for i in range(1,len(Need.index)+1):
            mid.append(i)
        Need.insert(0, 'id', mid)
        Need['长度'] = Need.apply(lambda x: round(int(x[1]) / int(x[0])/7.65/para*1000), axis=1)
        Need.insert(1, '需求类型', list_value)
        Need['已完成'] = ''
        Need['未完成'] = ''
        header_list = ['id', '需求类型','宽度','重量','长度','已完成','未完成']
        max_len = 0
        for i in range(len(prolist)):
            for k,v in prolist[i].items():
                now_len = len(v)
                if now_len > max_len:
                    max_len = now_len
        # print(max_len)
        for i in range(max_len):
            header_list.append('配刀'+str(i+1))
            header_list.append('米数'+str(i+1))
        for i in range(max_len*2):
            Need[str(i)] = ''
        Need.columns = header_list
        print(min_length)
        Need['宽度'] = Need['宽度'].apply(str) 
        for i in range(len(prolist)):
            if i == 0:
                for item,value in prolist[i].items():
                    sum = 0
                    for j in range(len(value)):
                        Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'), '配刀'+str(j+1)] = value[j][0]
                        Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'), '米数'+str(j+1)] = round(value[j][1])
                        sum = sum + round(value[j][1])
                    # print(item,sum)
                    Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'),'已完成'] = round(self.cal_weight(sum,int(float(item)),para))
                    Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'),'未完成'] = round(int(float(Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'), '重量'].values[0])) - self.cal_weight(sum,int(float(item)),para))
                    # if (Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'), '长度'].values[0] -sum) < min_length:
                    #     Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'),'未完成'] = 0
                    # else:
                    #     origin_length = Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'), '长度'].values[0]
                    #     if origin_length*(1-ner) <= sum and sum <= origin_length*(1+per):
                    #         Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'),'未完成'] = 0
                    #     else:
                    #         Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'),'未完成'] = int(float(Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '主项目'), '重量'].values[0])) - self.cal_weight(sum,int(float(item)),para)
            elif i == 1:
                for item,value in prolist[i].items():
                    sum = 0
                    for j in range(len(value)):
                        Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '次项目'), '配刀'+str(j+1)] = value[j][0]
                        Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '次项目'), '米数'+str(j+1)] = round(value[j][1])
                        sum = sum + round(value[j][1])
                    # print(item,sum)
                    Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '次项目'),'已完成'] = round(self.cal_weight(sum,int(float(item)),para))
                    Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '次项目'),'未完成'] = round(int(float(Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '次项目'), '重量'].values[0])) - self.cal_weight(sum,int(float(item)),para))
            else:
                for item,value in prolist[i].items():
                    sum = 0
                    for j in range(len(value)):
                        Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i+1)+'项目'), '配刀'+str(j+1)] = value[j][0]
                        Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i+1)+'项目'), '米数'+str(j+1)] = round(value[j][1])
                        sum = sum + round(value[j][1])
                    Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i+1)+'项目'),'已完成'] = round(self.cal_weight(sum,int(float(item)),para))
                    Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i+1)+'项目'),'未完成'] = round(int(float(Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i+1)+'项目'), '重量'].values[0])) - self.cal_weight(sum,int(float(item)),para))
                    # if (Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i)+'项目'), '长度'].values[0] -sum) < min_length:
                    #     Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i)+'项目'),'未完成'] = 0
                    # else:
                    #     origin_length = Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i)+'项目'), '长度'].values[0]
                    #     if origin_length*(1-ner) <= sum and sum <= origin_length*(1+per):
                    #         Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i)+'项目'),'未完成'] = 0
                    #     else:
                    #         Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i)+'项目'),'未完成'] = int(float(Need.loc[(Need['宽度'] == str(int(float(item)))) & (Need['需求类型'] == '第'+self.num_to_char(i)+'项目'), '重量'].values[0])) - self.cal_weight(sum,int(float(item)),para)
        # print(Need)
        self.Need_new = Need.copy()
        self.Need_save = Need.copy()
        self.Need_save = self.Need_save.drop(header_list[7:],axis=1)
        columns = self.Need_new.columns
        rows = self.Need_new.index

        self.tableWidget_6.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.tableWidget_6.setRowCount(len(rows))    # 设置表格的行数
        self.tableWidget_6.setColumnCount(len(columns))  # 设置表格的列数
        self.tableWidget_6.setHorizontalHeaderLabels(columns) # 设置表格的列标签
        self.tableWidget_6.verticalHeader().setHidden(True)
        
        # 将 EXCEL 文档的数据显示在表格中
        for i in range(len(rows)):
            for j in range(len(columns)):
                tstr = str(self.Need_new.iloc[i][j])
                data = QTableWidgetItem(tstr)
                self.tableWidget_6.setItem(i, j, data)
                self.tableWidget_6.item(i, j).setTextAlignment(Qt.AlignCenter or Qt.AlignHCenter)
        self.tableWidget_6.horizontalHeader().setStyleSheet("border-top:1px solid rgb(210, 210, 210);border-bottom:1px solid rgb(210, 210, 210);font:12pt '宋体';")
        self.tableWidget_6.resizeColumnsToContents() # 列宽随着内容调整
        self.tableWidget_6.resizeRowsToContents() # 行宽随着内容调整
        self.tableWidget_6.setAlternatingRowColors(True)  # 表格的颜色交错显示
        # return 1

    def Output(self, filename, message):
        wb = load_workbook(filename)
        sheets = wb.sheetnames       # 获取全部sheet
        ws = wb[sheets[0]]   # 切换  sheet
        row = ws.max_row    # 获取最大的 行数
        ws = wb.active

        for column, text in enumerate(message,start=1):
            ws.cell(column=column, row=row+1, value=text)

        wb.save(filename)

    def save_result(self):
        path = os.path.realpath(os.curdir) + '/结果'
        if not os.path.exists(path):
            os.makedirs(path)
        
        file_path,ok = QFileDialog.getSaveFileName(self.parent, "save file",path + '/结果.xlsx',"Excel files(*.xlsx)")
        if ok:
            self.Roll_file_new.drop(labels=['id'], axis=1,inplace = True)
            self.Need_save.drop(labels=['id'], axis=1,inplace = True)
            print(self.final_info)
            # self.Roll_file_new['指标'] = pd.Series(self.final_info).values
            write = pd.ExcelWriter(file_path)
            self.Roll_file_new.to_excel(write,sheet_name='原材料',index=False)
            self.Need_save.to_excel(write,sheet_name='需求满足情况',index=False)
            write.save()
            self.Output(file_path,self.final_info)

    def show_user_ui(self):
        self.User_MA = Ui_Dialog_User(self.parent)
        self.User_MA.exec()

    def login_out(self):
        self.parent.login_out()

